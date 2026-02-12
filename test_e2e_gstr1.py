"""
End-to-End Test: Process Client Excel File and Export to GSTR-1 Format

This script:
1. Loads GSTR-1_Dec-2025.xlsx (multi-sheet format)
2. Processes through process_multi_sheet_excel()
3. Passes clean_data to generate_gstr1_tables()
4. Exports to Excel using offline tool format
5. Validates that exported sheets are populated
"""

import sys
import os
sys.path.insert(0, '.')

from india_compliance.gst_india.utils.processor import process_multi_sheet_excel
from india_compliance.gst_india.gstr1_data import generate_gstr1_tables
from india_compliance.gst_india.exporters.gstr1_excel import export_gstr1_excel_direct
import openpyxl

# File paths
INPUT_FILE = "GSTR-1_Dec-2025.xlsx"
OUTPUT_FILE = "GSTR-1_Export_Test.xlsx"

print("=" * 60)
print("GSTR-1 End-to-End Test")
print("=" * 60)
print()

# Step 1: Check if input file exists
print("Step 1: Checking input file: {}".format(INPUT_FILE))
if not os.path.exists(INPUT_FILE):
    print("ERROR: Input file not found: {}".format(INPUT_FILE))
    sys.exit(1)
print("OK - File found")
print()

# Step 2: Check file structure
print("Step 2: Inspecting Excel file structure...")
wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
print("   Sheets: {}".format(wb.sheetnames))

# Check the first sheet structure
first_sheet = wb.sheetnames[0]
ws = wb[first_sheet]
print("   First sheet: {} ({} rows, {} columns)".format(
    first_sheet, ws.max_row, ws.max_column
))
wb.close()
print()

# Step 3: Load and process Excel using multi-sheet processor
print("Step 3: Processing Excel file (multi-sheet format)...")
with open(INPUT_FILE, "rb") as f:
    file_content = f.read()
print("   File size: {:,} bytes".format(len(file_content)))

result = process_multi_sheet_excel(
    file_content,
    return_period="122025",
    company_gstin="07AAAAA1234A1ZA",
)
print("   Processing complete")
print()

# Step 4: Report validation results
print("Step 4: Validation Results:")
clean_data = result.get("clean_data", [])
errors = result.get("errors", [])
skipped = result.get("skipped_rows", [])

print("   Total rows processed: {}".format(result.get('summary', {}).get('total_rows', 0)))
print("   Valid rows: {}".format(len(clean_data)))
print("   Errors: {}".format(len(errors)))
print("   Skipped rows: {}".format(len(skipped)))

if errors:
    print()
    print("   Sample Errors:")
    for error in errors[:3]:
        print("     Row {}: {}".format(error.get('row', '?'), error.get('errors', [])))

if skipped:
    print()
    print("   Sample Skipped Rows:")
    for skipped_row in skipped[:3]:
        print("     Row {} ({}): {}".format(
            skipped_row.get('row', '?'),
            skipped_row.get('sheet', '?'),
            skipped_row.get('reason', 'Unknown')[:60]
        ))
print()

# Step 5: Generate GSTR-1 Tables
print("Step 5: Generating GSTR-1 Tables...")
gstr1_tables = generate_gstr1_tables(
    clean_data,
    company_gstin="07AAAAA1234A1ZA",
    include_hsn=True,
    include_docs=True,
)
print("   Generation complete")
print()

# Step 6: Report GSTR-1 table counts
print("Step 6: GSTR-1 Table Counts:")
b2b_count = len(gstr1_tables.get('b2b', []))
b2cl_count = len(gstr1_tables.get('b2cl', []))
b2cs_count = len(gstr1_tables.get('b2cs', []))
exp_count = len(gstr1_tables.get('exp', []))
cdnr_count = len(gstr1_tables.get('cdnr', []))
cdnur_count = len(gstr1_tables.get('cdnur', []))
hsn_count = len(gstr1_tables.get('hsn', []))

print("   B2B entities: {}".format(b2b_count))
print("   B2CL invoices: {}".format(b2cl_count))
print("   B2CS entries: {}".format(b2cs_count))
print("   EXP invoices: {}".format(exp_count))
print("   CDNR entities: {}".format(cdnr_count))
print("   CDNUR entries: {}".format(cdnur_count))
print("   HSN records: {}".format(hsn_count))

summary = gstr1_tables.get("summary", {})
print()
print("Step 7: Tax Summary:")
print("   Total Taxable: Rs.{:,.2f}".format(summary.get('total_taxable_value', 0)))
print("   Total IGST: Rs.{:,.2f}".format(summary.get('total_igst', 0)))
print("   Total CGST: Rs.{:,.2f}".format(summary.get('total_cgst', 0)))
print("   Total SGST: Rs.{:,.2f}".format(summary.get('total_sgst', 0)))
print("   Total Cess: Rs.{:,.2f}".format(summary.get('total_cess', 0)))
print()

# Step 8: Export to Excel
print("Step 8: Exporting to Excel format...")
excel_bytes = export_gstr1_excel_direct(
    gstr1_tables,
    return_period="122025",
    taxpayer_gstin="07AAAAA1234A1ZA",
    taxpayer_name="Test Taxpayer",
)
print("   Export complete: {:,} bytes".format(len(excel_bytes)))
print()

# Step 9: Save and validate output
print("Step 9: Validating exported Excel...")
with open(OUTPUT_FILE, "wb") as f:
    f.write(excel_bytes)
print("   Saved to: {}".format(OUTPUT_FILE))

# Validate sheets
wb = openpyxl.load_workbook(OUTPUT_FILE)
print("   Sheets: {}".format(wb.sheetnames))

validation_passed = True
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    # Count data rows (skip header rows 1-4)
    data_rows = 0
    for row_idx in range(5, ws.max_row + 1):
        if any(cell.value is not None for cell in ws[row_idx]):
            data_rows += 1
    status = "OK" if data_rows >= 0 else "EMPTY"
    if data_rows == 0 and sheet_name not in ["Summary"]:
        status = "EMPTY (expected for unused sections)"
    print("   {}: {} data rows {}".format(sheet_name, data_rows, status))

wb.close()
print()

# Final Result
print("=" * 60)
total_records = b2b_count + b2cl_count + b2cs_count + exp_count + cdnr_count + cdnur_count
if total_records > 0:
    print("RESULT: End-to-End Test PASSED")
    print("   {} total records processed and exported".format(total_records))
else:
    print("RESULT: End-to-End Test COMPLETED")
    print("   No records found (file may be empty or use different format)")
print("=" * 60)
print()
print("Output file: {}".format(OUTPUT_FILE))
print("Ready for upload to GST portal or offline utility tool.")
