"""
Excel Upload Parser & Validator

This module processes uploaded Excel files containing sales data,
performs data cleaning, column mapping, and validation.

Features:
- Column header mapping with alias support
- GSTIN validation (15-character format)
- Place of Supply (POS) validation (2-digit state codes)
- Financial year date validation
- Tax rate validation (0, 5, 12, 18, 28)
- Tax breakup validation (CGST+SGST = IGST)
"""

from openpyxl import load_workbook
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
import re
import io

from india_compliance.gst_india.utils.logger import get_logger

# Initialize logger
logger = get_logger(__name__)


# Column header aliases mapping
# Maps common variations to standardized column names
COLUMN_ALIASES: Dict[str, List[str]] = {
    "gstin": [
        "GSTIN", "GSTIN/UIN", "GSTIN of Recipient", "GSTIN of B2B Recipient",
        "Recipient GSTIN", "GST Number", " gstin", " gstin"
    ],
    "invoice_number": [
        "Invoice Number", "Invoice No", "Invoice No.", "Invoice#",
        "Bill Number", "Voucher Number", "Invoice"
    ],
    "invoice_date": [
        "Invoice Date", "Invoice Date ", "Date", "Invoice Dt",
        "Bill Date", "Transaction Date", "Document Date"
    ],
    "invoice_value": [
        "Invoice Value", "Total Invoice Value", "Invoice Amount",
        "Bill Value", "Total Amount", "Grand Total", "Total"
    ],
    "place_of_supply": [
        "Place Of Supply", "POS", "Place of Supply", "State",
        "Supply State", "Destination State"
    ],
    "taxable_value": [
        "Taxable Value", "Taxable Amount", "Net Amount",
        "Taxable", "Amount", "Taxable Value "
    ],
    "rate": [
        "Rate", "Tax Rate", "GST Rate", "Rate %", "Percentage",
        "Tax Percentage", "Rate(%)"
    ],
    "cgst": [
        "CGST", "CGST Amount", "Central GST", "Central Tax",
        "CGST ", " cgst"
    ],
    "sgst": [
        "SGST", "SGST Amount", "State GST", "State Tax",
        "SGST ", " sgst"
    ],
    "igst": [
        "IGST", "IGST Amount", "Integrated GST", "Integrated Tax",
        "IGST ", " igst"
    ],
    "cess": [
        "CESS", "CESS Amount", " cess", " CESS"
    ],
    "customer_name": [
        "Customer Name", "Recipient Name", "Party Name",
        "Name of Recipient", "Consignee Name", "Buyer Name"
    ],
    "invoice_type": [
        "Invoice Type", "Type", "Document Type"
    ],
}

# Standard tax rates
VALID_TAX_RATES = {0, 0.1, 0.25, 1, 1.5, 3, 5, 6, 7.5, 12, 18, 28, 40}

# Valid state codes (2-digit)
VALID_STATE_CODES = {
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
    "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
    "21", "22", "23", "24", "25", "26", "27", "28", "29", "30",
    "31", "32", "33", "34", "35", "36", "37", "38", "96", "97"
}

# GSTIN regex pattern
GSTIN_PATTERN = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}[Z1-9ABD-J]{1}[0-9A-Z]{1}$")

# Current financial year
def get_current_financial_year() -> Tuple[int, int]:
    """Get the current financial year (start, end)."""
    today = date.today()
    if today.month >= 4:
        start_year = today.year
        end_year = today.year + 1
    else:
        start_year = today.year - 1
        end_year = today.year
    return (start_year, end_year)


def validate_gstin(gstin: str) -> Tuple[bool, str]:
    """
    Validate GSTIN format (15 characters).
    
    Args:
        gstin: GSTIN string to validate
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if not gstin:
        return (False, "GSTIN is required")
    
    # Clean GSTIN (remove spaces)
    gstin = str(gstin).strip().upper().replace(" ", "")
    
    # Check length
    if len(gstin) != 15:
        return (False, f"GSTIN must be 15 characters (got {len(gstin)})")
    
    # Check pattern
    if not GSTIN_PATTERN.match(gstin):
        return (False, "Invalid GSTIN format")
    
    return (True, "")


def validate_pos(pos: str) -> Tuple[bool, str]:
    """
    Validate Place of Supply (2-digit state code).
    
    Args:
        pos: Place of Supply string to validate
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if not pos:
        return (False, "Place of Supply is required")
    
    # Clean POS
    pos = str(pos).strip()
    
    # Try to extract 2-digit code
    code_match = re.match(r"^(\d{2})", pos)
    if code_match:
        code = code_match.group(1)
        if code in VALID_STATE_CODES:
            return (True, "")
        return (False, f"Invalid state code: {code}")
    
    # Check if it's a state name
    from india_compliance.gst_india.constants import STATE_NUMBERS
    pos_normalized = pos.title()
    if pos_normalized in STATE_NUMBERS:
        return (True, "")
    
    return (False, f"Invalid Place of Supply: {pos}")


def validate_invoice_date(invoice_date: Any) -> Tuple[bool, str, Optional[datetime]]:
    """
    Validate invoice date is within current or previous financial year.
    
    Args:
        invoice_date: Date string or datetime object
        
    Returns:
        Tuple of (is_valid, error_message, parsed_date)
    """
    if not invoice_date:
        return (False, "Invoice date is required", None)
    
    # Parse date
    parsed_date = None
    if isinstance(invoice_date, (date, datetime)):
        parsed_date = invoice_date if isinstance(invoice_date, datetime) else datetime.combine(invoice_date, datetime.min.time())
    else:
        # Try various date formats
        date_formats = [
            "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
            "%d/%m/%y", "%d-%m-%y", "%Y/%m/%d",
            "%d %B %Y", "%d %b %Y", "%B %d, %Y"
        ]
        
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(str(invoice_date).strip(), fmt)
                break
            except ValueError:
                continue
        
        if not parsed_date:
            return (False, f"Invalid date format: {invoice_date}", None)
    
    # Check financial year
    start_year, end_year = get_current_financial_year()
    fy_start = date(start_year, 4, 1)
    fy_end = date(end_year, 3, 31)
    prev_fy_start = date(start_year - 1, 4, 1)
    prev_fy_end = date(end_year - 1, 3, 31)
    
    invoice_date_only = parsed_date.date()
    
    # Check current or previous financial year
    if (fy_start <= invoice_date_only <= fy_end) or (prev_fy_start <= invoice_date_only <= prev_fy_end):
        return (True, "", parsed_date)
    
    return (False, f"Invoice date {invoice_date_only} is outside current/previous financial year", None)


def validate_tax_rate(tax_rate: Any) -> Tuple[bool, str, float]:
    """
    Validate tax rate is one of [0, 5, 12, 18, 28].
    
    Args:
        tax_rate: Tax rate value
        
    Returns:
        Tuple of (is_valid, error_message, parsed_rate)
    """
    if tax_rate is None or tax_rate == "":
        return (False, "Tax rate is required", 0.0)
    
    # Parse rate
    try:
        rate = float(str(tax_rate).strip().replace("%", ""))
    except ValueError:
        return (False, f"Invalid tax rate: {tax_rate}", 0.0)
    
    # Round rate for comparison
    rate = round(rate, 2)
    
    # Check valid rates
    if rate in VALID_TAX_RATES:
        return (True, "", rate)
    
    return (False, f"Invalid tax rate: {rate} (must be 0, 5, 12, 18, or 28)", rate)


def validate_tax_breakup(
    taxable_value: float,
    rate: float,
    cgst: Optional[float],
    sgst: Optional[float],
    igst: Optional[float]
) -> Tuple[bool, str]:
    """
    Validate tax breakup matches taxable value * rate.
    
    Args:
        taxable_value: Taxable amount
        rate: Tax rate percentage
        cgst: CGST amount
        sgst: SGST amount
        igst: IGST amount
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if taxable_value is None or rate is None:
        return (True, "")  # Skip if values missing
    
    # Calculate expected tax
    expected_total_tax = float(taxable_value) * rate / 100
    
    # Round for comparison
    expected_total_tax = round(expected_total_tax, 2)
    
    # Get actual tax
    if igst is not None and igst > 0:
        actual_tax = float(igst) if igst else 0
    elif cgst is not None and sgst is not None:
        actual_tax = float(cgst or 0) + float(sgst or 0)
    else:
        # No tax info provided, skip validation
        return (True, "")
    
    actual_tax = round(float(actual_tax), 2)
    
    # Check if tax amounts match (allow small difference for rounding)
    tolerance = 0.05  # 5 paise tolerance
    if abs(expected_total_tax - actual_tax) > tolerance:
        return (
            False,
            f"Tax mismatch: Expected {expected_total_tax:.2f}, Got {actual_tax:.2f} "
            f"(Taxable: {taxable_value}, Rate: {rate}%)"
        )
    
    return (True, "")


def find_column(headers: List[str], column_type: str) -> Optional[str]:
    """
    Find the actual column header that matches a column type.
    
    Args:
        headers: List of column headers from Excel
        column_type: Type of column to find (e.g., "gstin")
        
    Returns:
        Matched column header or None
    """
    aliases = COLUMN_ALIASES.get(column_type, [])
    
    for alias in aliases:
        for header in headers:
            if header.strip().lower() == alias.lower():
                return header
    
    # Try partial matching
    for header in headers:
        header_lower = header.strip().lower()
        for alias in aliases:
            alias_lower = alias.lower()
            if alias_lower in header_lower or header_lower in alias_lower:
                return header
    
    return None


def clean_value(value: Any) -> Any:
    """
    Clean and normalize a value from Excel.
    
    Args:
        value: Raw value from Excel
        
    Returns:
        Cleaned value
    """
    if value is None:
        return None
    
    if isinstance(value, str):
        return value.strip() or None
    
    return value


def map_columns(row: Dict[str, Any], column_mapping: Dict[str, str]) -> Dict[str, Any]:
    """
    Map row data to standardized column names.
    
    Args:
        row: Raw row data with column names as keys
        column_mapping: Mapping from standard names to actual column names
        
    Returns:
        Row data with standardized column names
    """
    mapped_row = {}
    
    for std_name, actual_name in column_mapping.items():
        if actual_name and actual_name in row:
            value = clean_value(row[actual_name])
            mapped_row[std_name] = value
    
    return mapped_row


def process_excel(
    file_content: bytes,
    expected_columns: Optional[List[str]] = None
) -> Dict[str, Any]:
    """
    Process an Excel file and validate all rows.
    
    Args:
        file_content: Raw bytes of the Excel file
        expected_columns: Optional list of expected column types to validate
        
    Returns:
        Dictionary with:
        - "clean_data": List of valid rows
        - "errors": List of validation errors
        - "summary": Processing summary
    """
    clean_data = []
    errors = []
    
    try:
        # Load workbook from bytes
        workbook = load_workbook(io.BytesIO(file_content))
        worksheet = workbook.active
        
        # Get headers
        headers = []
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)):
            if cell:
                headers.append(str(cell))
        
        logger.info(f"Excel file loaded with {len(headers)} columns: {headers}")
        
        # Build column mapping
        column_mapping = {}
        if expected_columns is None:
            expected_columns = list(COLUMN_ALIASES.keys())
        
        for col_type in expected_columns:
            found = find_column(headers, col_type)
            if found:
                column_mapping[col_type] = found
        
        logger.info(f"Column mapping: {column_mapping}")
        
        # Process rows (skip header row)
        row_number = 1  # Start from 1 (after header)
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_number += 1
            row_data = {}
            
            # Create row dictionary
            for idx, cell_value in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = cell_value
            
            # Map columns
            mapped_row = map_columns(row_data, column_mapping)
            
            # Validate row
            row_errors = []
            
            # Validate GSTIN if present
            if mapped_row.get("gstin"):
                is_valid, error = validate_gstin(mapped_row["gstin"])
                if not is_valid:
                    row_errors.append({"field": "gstin", "error": error})
            
            # Validate POS if present
            if mapped_row.get("place_of_supply"):
                is_valid, error = validate_pos(mapped_row["place_of_supply"])
                if not is_valid:
                    row_errors.append({"field": "place_of_supply", "error": error})
            
            # Validate invoice date if present
            if mapped_row.get("invoice_date"):
                is_valid, error, parsed_date = validate_invoice_date(mapped_row["invoice_date"])
                if not is_valid:
                    row_errors.append({"field": "invoice_date", "error": error})
                else:
                    mapped_row["invoice_date"] = parsed_date
            
            # Validate tax rate if present
            if mapped_row.get("rate"):
                is_valid, error, rate = validate_tax_rate(mapped_row["rate"])
                if not is_valid:
                    row_errors.append({"field": "rate", "error": error})
                else:
                    mapped_row["rate"] = rate
            
            # Validate tax breakup if taxable value and rate present
            if mapped_row.get("taxable_value") is not None and mapped_row.get("rate") is not None:
                is_valid, error = validate_tax_breakup(
                    mapped_row["taxable_value"],
                    mapped_row["rate"],
                    mapped_row.get("cgst"),
                    mapped_row.get("sgst"),
                    mapped_row.get("igst")
                )
                if not is_valid:
                    row_errors.append({"field": "tax_breakup", "error": error})
            
            # Add row to results
            if row_errors:
                errors.append({
                    "row": row_number,
                    "errors": row_errors,
                    "data": mapped_row
                })
            else:
                # Clean numeric values
                for key in ["taxable_value", "invoice_value", "cgst", "sgst", "igst", "cess"]:
                    if mapped_row.get(key) is not None:
                        try:
                            mapped_row[key] = float(mapped_row[key])
                        except (ValueError, TypeError):
                            pass
                
                clean_data.append(mapped_row)
        
        # Log summary
        logger.info(
            f"Excel processing completed: {len(clean_data)} valid rows, {len(errors)} rows with errors"
        )
        
        # Build summary
        def safe_float(value, default=0.0):
            return float(value) if value is not None else default
        
        total_taxable = sum(safe_float(row.get("taxable_value", 0)) for row in clean_data)
        total_igst = sum(safe_float(row.get("igst", 0)) for row in clean_data)
        total_cgst = sum(safe_float(row.get("cgst", 0)) for row in clean_data)
        total_sgst = sum(safe_float(row.get("sgst", 0)) for row in clean_data)
        total_cess = sum(safe_float(row.get("cess", 0)) for row in clean_data)
        
        summary = {
            "total_rows": len(clean_data) + len(errors),
            "valid_rows": len(clean_data),
            "error_rows": len(errors),
            "total_taxable_value": round(total_taxable, 2),
            "total_igst": round(total_igst, 2),
            "total_cgst": round(total_cgst, 2),
            "total_sgst": round(total_sgst, 2),
            "total_cess": round(total_cess, 2),
        }
        
        return {
            "clean_data": clean_data,
            "errors": errors,
            "summary": summary
        }
    
    except Exception as e:
        logger.exception(f"Error processing Excel file: {str(e)}")
        return {
            "clean_data": [],
            "errors": [{"row": 0, "error": f"File processing error: {str(e)}"}],
            "summary": {"total_rows": 0, "valid_rows": 0, "error_rows": 1}
        }


def process_gstr1_excel(file_content: Any) -> Dict[str, Any]:
    """
    Process Excel file for GSTR-1 format.
    
    Args:
        file_content: Raw bytes of the Excel file OR file path (str)
        
    Returns:
        Dictionary with GSTR-1 structured data
    """
    # If file_content is a string (file path), read the file
    if isinstance(file_content, str):
        with open(file_content, "rb") as f:
            file_content = f.read()
    
    result = process_excel(file_content)
    
    if result["errors"]:
        logger.warning(f"Found {len(result['errors'])} validation errors in Excel file")
    
    # Transform to GSTR-1 format
    gstr1_data = {
        "b2b": [],
        "b2cl": [],
        "b2cs": [],
        "export": [],
        "cdnr": [],
        "cdnur": [],
        "summary": result["summary"]
    }
    
    for row in result["clean_data"]:
        # Classify row based on GSTIN and POS
        gstin = row.get("gstin", "")
        
        if not gstin:
            # Unregistered - classify as B2CS or Export
            pos = str(row.get("place_of_supply", ""))
            
            # Check if export (overseas)
            if "overseas" in pos.lower() or "export" in pos.lower():
                gstr1_data["export"].append({
                    "invoice_number": row.get("invoice_number"),
                    "invoice_date": row.get("invoice_date"),
                    "invoice_value": row.get("invoice_value"),
                    "shipping_port": row.get("place_of_supply"),
                    "taxable_value": row.get("taxable_value"),
                    "rate": row.get("rate"),
                    "igst": row.get("igst", 0),
                    "cess": row.get("cess", 0),
                })
            else:
                # B2CS
                gstr1_data["b2cs"].append({
                    "invoice_date": row.get("invoice_date"),
                    "invoice_value": row.get("invoice_value"),
                    "place_of_supply": row.get("place_of_supply"),
                    "taxable_value": row.get("taxable_value"),
                    "rate": row.get("rate"),
                    "igst": row.get("igst", 0),
                    "cess": row.get("cess", 0),
                })
        else:
            # Registered - classify as B2B or B2CL
            pos = str(row.get("place_of_supply", ""))
            state_code = pos[:2] if len(pos) >= 2 else ""
            
            # Check if inter-state (different state)
            # For B2CL, IGST should be applied
            if row.get("igst", 0) > 0:
                gstr1_data["b2cl"].append({
                    "invoice_number": row.get("invoice_number"),
                    "invoice_date": row.get("invoice_date"),
                    "invoice_value": row.get("invoice_value"),
                    "place_of_supply": row.get("place_of_supply"),
                    "taxable_value": row.get("taxable_value"),
                    "rate": row.get("rate"),
                    "igst": row.get("igst", 0),
                    "cess": row.get("cess", 0),
                })
            else:
                # B2B
                gstr1_data["b2b"].append({
                    "gstin": row.get("gstin"),
                    "customer_name": row.get("customer_name"),
                    "invoice_number": row.get("invoice_number"),
                    "invoice_date": row.get("invoice_date"),
                    "invoice_value": row.get("invoice_value"),
                    "place_of_supply": row.get("place_of_supply"),
                    "taxable_value": row.get("taxable_value"),
                    "rate": row.get("rate"),
                    "cgst": row.get("cgst", 0),
                    "sgst": row.get("sgst", 0),
                    "cess": row.get("cess", 0),
                })
    
    logger.info(
        f"GSTR-1 data extracted: "
        f"{len(gstr1_data['b2b'])} B2B, "
        f"{len(gstr1_data['b2cl'])} B2CL, "
        f"{len(gstr1_data['b2cs'])} B2CS, "
        f"{len(gstr1_data['export'])} Export"
    )
    
    return {
        **gstr1_data,
        "clean_data": result["clean_data"],
        "errors": result["errors"],
        "validation_summary": {
            "errors": result["errors"],
            "warnings": [],
        }
    }
