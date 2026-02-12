from openpyxl import load_workbook
wb = load_workbook('gstr1_122025.xlsx')

for sheet_name in ['B2CL', 'B2CS', 'EXP', 'CDNR', 'HSN']:
    ws = wb[sheet_name]
    print(f'=== {sheet_name} ===')
    print(f'Rows: {ws.max_row}, Cols: {ws.max_column}')
    for row in ws.iter_rows(min_row=4, max_row=min(ws.max_row, 10), values_only=True):
        print([str(c)[:15] if c else '' for c in row])
    print()
