#!/usr/bin/env python
"""Verify generated Excel file structure."""

from openpyxl import load_workbook
import io

# Load the generated Excel file
with open("test_gstr1_output.xlsx", "rb") as f:
    content = f.read()

workbook = load_workbook(io.BytesIO(content), data_only=True)

print("=== Generated GSTR-1 Excel Structure ===\n")

for sheet_name in workbook.sheetnames:
    worksheet = workbook[sheet_name]
    print(f"Sheet: {sheet_name}")
    print(f"  Max rows: {worksheet.max_row}, Max cols: {worksheet.max_column}")
    
    # Print headers (row 4 = index 3 for GSTR-1 format)
    if worksheet.max_row >= 4:
        print("  Headers (row 4):")
        headers = []
        for col in range(1, min(worksheet.max_column + 1, 15)):
            cell_value = worksheet.cell(row=4, column=col).value
            headers.append(str(cell_value) if cell_value else "")
            print(f"    Col {col}: {cell_value}")
    print()

workbook.close()
print("=== Excel verification complete ===")
