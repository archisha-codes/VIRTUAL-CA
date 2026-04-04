"""
Excel Upload Parser & Validator

This module processes uploaded Excel files containing sales data,
performs data cleaning, column mapping, and validation.

Features:
- Column header mapping with alias support
- GSTIN validation (15-character format)
- Place of Supply (POS) validation (2-digit state codes)
- Financial year date validation
- Tax rate validation (0, 5, 12, 18, 28)
- Tax breakup validation (CGST+SGST = IGST)
- Memory-efficient streaming for large files (10,000+ rows)
- Timer logs and warnings for processing >5 seconds
"""

from openpyxl import load_workbook
from typing import Dict, Any, List, Optional, Tuple, Generator
from datetime import datetime, date
from decimal import Decimal, ROUND_HALF_UP
import re
import io
import time

# Try to import pandas for skiprows support (Tally exports)
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None

from india_compliance.gst_india.utils.logger import get_logger

# Initialize logger
logger = get_logger(__name__)

# Processing timeout threshold (seconds)
PROCESSING_TIMEOUT_THRESHOLD = 5


def log_processing_time(func):
    """Decorator to log processing time and warn if exceeding threshold."""
    def wrapper(*args, **kwargs):
        start_time = time.time()
        result = func(*args, **kwargs)
        elapsed = time.time() - start_time
        
        if elapsed > PROCESSING_TIMEOUT_THRESHOLD:
            logger.warning(
                f"{func.__name__} took {elapsed:.2f}s (> {PROCESSING_TIMEOUT_THRESHOLD}s threshold). "
                f"Consider optimizing for large datasets."
            )
        else:
            logger.debug(f"{func.__name__} completed in {elapsed:.2f}s")
        
        return result
    return wrapper


# Column header aliases mapping
# Maps common variations to standardized column names
COLUMN_ALIASES: Dict[str, List[str]] = {
    "gstin": [
        "GSTIN", "GSTIN/UIN", "GSTIN of Recipient", "GSTIN of B2B Recipient",
        "Recipient GSTIN", "GST Number", " gstin", " gstin",
        "Tax Number 3", "TaxNumber3", "GSTIN No", "GST No"
    ],
    "invoice_number": [
        "Invoice Number", "Invoice No", "Invoice No.", "Invoice#",
        "Bill Number", "Voucher Number", "Invoice",
        "Document Number", "Doc No", "Voucher No", "Voucher Number"
    ],
    "invoice_date": [
        "Invoice Date", "Invoice Date ", "Date", "Invoice Dt",
        "Bill Date", "Transaction Date", "Document Date",
        "Voucher Date", "Doc Date"
    ],
    "invoice_value": [
        "Invoice Value", "Total Invoice Value", "Invoice Amount",
        "Bill Value", "Total Amount", "Grand Total", "Total",
        "Voucher Value", "Doc Value", "Basic Amount", "Basic"
    ],
    "place_of_supply": [
        "Place Of Supply", "POS", "Place of Supply", "State",
        "Supply State", "Destination State",
        "Tax Country", "Country"
    ],
    "taxable_value": [
        "Taxable Value", "Taxable Amount", "Net Amount",
        "Taxable", "Amount", "Taxable Value ",
        "Taxable Value (₹)", "Taxable Amount (₹)"
    ],
    "rate": [
        "Rate", "Tax Rate", "GST Rate", "Rate %", "Percentage",
        "Tax Percentage", "Rate(%)", "Rate ( % )", "Tax %"
    ],
    "cgst": [
        "CGST", "CGST Amount", "Central GST", "Central Tax",
        "CGST ", " cgst", "CGST (₹)"
    ],
    "sgst": [
        "SGST", "SGST Amount", "State GST", "State Tax",
        "SGST ", " sgst", "SGST (₹)"
    ],
    "igst": [
        "IGST", "IGST Amount", "Integrated GST", "Integrated Tax",
        "IGST ", " igst", "IGST (₹)"
    ],
    "cess": [
        "CESS", "CESS Amount", " cess", " CESS", "Cess (₹)"
    ],
    "customer_name": [
        "Customer Name", "Receiver Name", "Recipient Name", "Party Name",
        "Name of Recipient", "Consignee Name", "Buyer Name",
        "Ledger Name", "Party"
    ],
    "invoice_type": [
        "Invoice Type", "Type", "Document Type", "Type of invoice",
        "Voucher Type Name", "Voucher Type"
    ],
    # E-invoice related fields
    "irn": [
        "IRN", "Invoice Reference Number", "e-Invoice IRN",
        "IRN No", "IRN Number", "GST IRN"
    ],
    "ack_no": [
        "Ack No", "Ack Number", "Acknowledgement No",
        "Acknowledgement Number", "Ack. No.", "Ack No."
    ],
    "ack_date": [
        "Ack Date", "Ack Date ", "Acknowledgement Date",
        "Ack. Date", "Date of Acknowledgement"
    ],
    # Additional fields
    "hsn_code": [
        "HSN", "HSN Code", "HSN/SAC", "SAC Code", "GST HSN",
        "HSN Code ", "HSN/SAC Code"
    ],
    "description": [
        "Description", "Item Description", "Product Description",
        "Goods/Services Description", "Particulars"
    ],
    "quantity": [
        "Quantity", "Qty", "Qty.", "Quantity ", "Total Quantity"
    ],
    "uom": [
        "UOM", "Unit", "Unit of Measure", "UQC", "UoM"
    ],
}

# Tally-specific column name mappings (exact column names from Tally exports)
TALLY_COLUMN_MAPPINGS = {
    # Tally often uses these exact column names
    "Tax Number 3": "gstin",
    "Voucher Number": "invoice_number",
    "Voucher Date": "invoice_date",
    "Voucher Value": "invoice_value",
    "Ledger Name": "customer_name",
    "Voucher Type Name": "invoice_type",
    "Basic Amount": "taxable_value",
    "Basic": "taxable_value",
    "Tax Type": "rate",
    "Tax %": "rate",
    "Tax Amount": "igst",
    "CGST Amount": "cgst",
    "SGST Amount": "sgst",
    "IGST Amount": "igst",
    "Cess Amount": "cess",
    "Tax Country": "place_of_supply",
    "Country of Origin": "place_of_supply",
    "Port Code": "port_code",
    "Shipping Bill No": "shipping_bill_number",
    "Shipping Bill Date": "shipping_bill_date",
    "Export Type": "export_type",
    "Nature of Transaction": "transaction_type",
}

# Standard tax rates
VALID_TAX_RATES = {0, 0.1, 0.25, 1, 1.5, 3, 5, 6, 7.5, 12, 18, 28, 40}

# Valid state codes (2-digit)
VALID_STATE_CODES = {
    "01", "02", "03", "04", "05", "06", "07", "08", "09", "10",
    "11", "12", "13", "14", "15", "16", "17", "18", "19", "20",
    "21", "22", "23", "24", "25", "26", "27", "28", "29", "30",
    "31", "32", "33", "34", "35", "36", "37", "38", "96", "97"
}

# State code to POS mapping (for IRN-generated data)
STATE_CODE_TO_POS = {
    "KA": "29",  # Karnataka
    "MH": "27",  # Maharashtra
    "DL": "07",  # Delhi
    "TN": "33",  # Tamil Nadu
    "TG": "36",  # Telangana
    "GJ": "24",  # Gujarat
    "RJ": "08",  # Rajasthan
    "UP": "09",  # Uttar Pradesh
    "MP": "23",  # Madhya Pradesh
    "WB": "19",  # West Bengal
    "BR": "10",  # Bihar
    "KL": "32",  # Kerala
    "KE": "32",  # Kerala (alternate)
    "AP": "28",  # Andhra Pradesh
    "JK": "01",  # Jammu and Kashmir
    "HP": "02",  # Himachal Pradesh
    "PN": "30",  # Punjab
    "CH": "04",  # Chandigarh
    "OR": "21",  # Odisha
    "CT": "22",  # Chhattisgarh
    "HR": "06",  # Haryana
    "UR": "96",  # Outside India (Export)
}

# GSTIN regex pattern
GSTIN_PATTERN = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}[Z1-9ABD-J]{1}[0-9A-Z]{1}$")

# Current financial year
def get_current_financial_year() -> Tuple[int, int]:
    """Get the current financial year (start, end)."""
    today = date.today()
    if today.month >= 4:
        start_year = today.year
        end_year = today.year + 1
    else:
        start_year = today.year - 1
        end_year = today.year
    return (start_year, end_year)


def validate_gstin(gstin: str, is_mandatory: bool = True) -> Tuple[bool, str]:
    """
    Validate GSTIN format (15 characters, PAN format).
    
    Args:
        gstin: GSTIN string to validate
        is_mandatory: Whether GSTIN is mandatory for this row
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    # Check if GSTIN is required
    if not gstin or not gstin.strip():
        if is_mandatory:
            return (False, "GSTIN is required")
        return (True, "")  # Empty GSTIN is allowed for B2C
    
    # Clean GSTIN (remove spaces)
    gstin = str(gstin).strip().upper().replace(" ", "")
    
    # Check length
    if len(gstin) != 15:
        return (False, f"GSTIN must be 15 characters (got {len(gstin)})")
    
    # Check pattern (2 digits + 5 uppercase letters + 4 digits + 1 uppercase letter + 1 alphanumeric + 1 check digit + 1 alphanumeric)
    if not GSTIN_PATTERN.match(gstin):
        return (False, "Invalid GSTIN format (must match PAN-based format)")
    
    return (True, "")


def validate_pos(pos: str) -> Tuple[bool, str]:
    """
    Validate Place of Supply (2-digit state code).
    
    Args:
        pos: Place of Supply string to validate
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if not pos:
        return (False, "Place of Supply is required")
    
    # Clean POS
    pos = str(pos).strip()
    
    # Try to extract 2-digit code
    code_match = re.match(r"^(\d{2})", pos)
    if code_match:
        code = code_match.group(1)
        if code in VALID_STATE_CODES:
            return (True, "")
        return (False, f"Invalid state code: {code}")
    
    # Check if it's a state name
    from india_compliance.gst_india.constants import STATE_NUMBERS
    pos_normalized = pos.title()
    if pos_normalized in STATE_NUMBERS:
        return (True, "")
    
    return (False, f"Invalid Place of Supply: {pos}")


def validate_invoice_date(invoice_date: Any) -> Tuple[bool, str, Optional[datetime]]:
    """
    Validate invoice date is within current or previous financial year.
    
    Args:
        invoice_date: Date string or datetime object
        
    Returns:
        Tuple of (is_valid, error_message, parsed_date)
    """
    if not invoice_date:
        return (False, "Invoice date is required", None)
    
    # Parse date
    parsed_date = None
    if isinstance(invoice_date, (date, datetime)):
        parsed_date = invoice_date if isinstance(invoice_date, datetime) else datetime.combine(invoice_date, datetime.min.time())
    else:
        # Try various date formats
        date_formats = [
            "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
            "%d/%m/%y", "%d-%m-%y", "%Y/%m/%d",
            "%d %B %Y", "%d %b %Y", "%B %d, %Y"
        ]
        
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(str(invoice_date).strip(), fmt)
                break
            except ValueError:
                continue
        
        if not parsed_date:
            return (False, f"Invalid date format: {invoice_date}", None)
    
    # Check financial year
    start_year, end_year = get_current_financial_year()
    fy_start = date(start_year, 4, 1)
    fy_end = date(end_year, 3, 31)
    prev_fy_start = date(start_year - 1, 4, 1)
    prev_fy_end = date(end_year - 1, 3, 31)
    
    invoice_date_only = parsed_date.date()
    
    # Check current or previous financial year
    if (fy_start <= invoice_date_only <= fy_end) or (prev_fy_start <= invoice_date_only <= prev_fy_end):
        return (True, "", parsed_date)
    
    return (False, f"Invoice date {invoice_date_only} is outside current/previous financial year", None)


def validate_tax_rate(tax_rate: Any) -> Tuple[bool, str, float]:
    """
    Validate tax rate is one of [0, 5, 12, 18, 28].
    
    Args:
        tax_rate: Tax rate value
        
    Returns:
        Tuple of (is_valid, error_message, parsed_rate)
    """
    if tax_rate is None or tax_rate == "":
        return (False, "Tax rate is required", 0.0)
    
    # Parse rate
    try:
        rate = float(str(tax_rate).strip().replace("%", ""))
    except ValueError:
        return (False, f"Invalid tax rate: {tax_rate}", 0.0)
    
    # Round rate for comparison
    rate = round(rate, 2)
    
    # Check valid rates
    if rate in VALID_TAX_RATES:
        return (True, "", rate)
    
    return (False, f"Invalid tax rate: {rate} (must be 0, 5, 12, 18, or 28)", rate)


def validate_tax_breakup(
    taxable_value: float,
    rate: float,
    cgst: Optional[float],
    sgst: Optional[float],
    igst: Optional[float],
    pos: str = ""
) -> Tuple[bool, str]:
    """
    Validate tax breakup matches taxable value * rate.
    Validates CGST+SGST = IGST for intra/inter-state logic.
    
    Args:
        taxable_value: Taxable amount
        rate: Tax rate percentage
        cgst: CGST amount
        sgst: SGST amount
        igst: IGST amount
        pos: Place of Supply (for determining inter/intra state)
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if taxable_value is None or rate is None:
        return (True, "")  # Skip if values missing
    
    # Calculate expected tax
    expected_total_tax = float(taxable_value) * rate / 100
    
    # Round for comparison
    expected_total_tax = round(expected_total_tax, 2)
    
    # Get actual tax values
    igst_val = float(igst) if igst is not None else 0
    cgst_val = float(cgst) if cgst is not None else 0
    sgst_val = float(sgst) if sgst is not None else 0
    
    # Determine if inter-state or intra-state
    is_inter_state = False
    if pos:
        pos_code = str(pos)[:2] if len(str(pos)) >= 2 else ""
        # Inter-state if state code is different from company's state
        # For now, just check if IGST is provided
        is_inter_state = igst_val > 0
    
    if is_inter_state:
        # Inter-state: Should have IGST, CGST and SGST should be 0
        if cgst_val > 0 or sgst_val > 0:
            if igst_val > 0:
                return (False, "For inter-state supply, CGST and SGST should be 0 when IGST is applicable")
        
        # Check IGST matches expected
        if igst_val > 0:
            if abs(igst_val - expected_total_tax) > 0.05:
                return (
                    False,
                    f"IGST mismatch: Expected {expected_total_tax:.2f}, Got {igst_val:.2f}"
                )
    else:
        # Intra-state: Should have CGST + SGST = IGST, IGST should be 0
        if igst_val > 0:
            return (False, "For intra-state supply, IGST should be 0 (use CGST+SGST)")
        
        # Check CGST + SGST matches expected
        actual_tax = cgst_val + sgst_val
        if abs(actual_tax - expected_total_tax) > 0.05:
            return (
                False,
                f"CGST+SGST mismatch: Expected {expected_total_tax:.2f}, Got {actual_tax:.2f} "
                f"(CGST: {cgst_val:.2f}, SGST: {sgst_val:.2f})"
            )
    
    return (True, "")


def validate_return_period(invoice_date: Any, return_period: str) -> Tuple[bool, str]:
    """
    Validate invoice date falls within return period month.
    
    Args:
        invoice_date: Invoice date (datetime or string)
        return_period: Return period in MMYYYY format
        
    Returns:
        Tuple of (is_valid, error_message)
    """
    if not return_period or len(return_period) != 6:
        return (True, "")  # Skip if return_period not provided
    
    # Parse return period
    try:
        return_month = int(return_period[:2])
        return_year = int(return_period[2:])
    except ValueError:
        return (True, "")  # Skip validation if format is wrong
    
    # Parse invoice date
    parsed_date = None
    if isinstance(invoice_date, (date, datetime)):
        parsed_date = invoice_date if isinstance(invoice_date, datetime) else datetime.combine(invoice_date, datetime.min.time())
    elif isinstance(invoice_date, str):
        date_formats = ["%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y"]
        for fmt in date_formats:
            try:
                parsed_date = datetime.strptime(invoice_date.strip(), fmt)
                break
            except ValueError:
                continue
    
    if not parsed_date:
        return (True, "")  # Skip if can't parse date
    
    # Check if invoice date falls within return period month
    if parsed_date.month == return_month and parsed_date.year == return_year:
        return (True, "")
    
    # Allow invoices from previous month of same financial year
    if parsed_date.month == return_month - 1 or (return_month == 1 and parsed_date.month == 12 and parsed_date.year == return_year - 1):
        return (True, "")
    
    return (
        False,
        f"Invoice date {parsed_date.strftime('%d/%m/%Y')} is outside return period {return_period}"
    )


def validate_row_for_gstr1(row: Dict[str, Any], row_number: int, sheet_name: str = "") -> Tuple[bool, List[Dict[str, str]], Dict[str, Any]]:
    """
    Validate a single row for GSTR-1 requirements.
    
    Checks all required fields and raises errors if critical fields are missing.
    Validates number parsing from strings.
    
    Args:
        row: Row data dictionary
        row_number: Row number for error reporting
        sheet_name: Sheet name for error context
        
    Returns:
        Tuple of (is_valid, errors, cleaned_row)
    """
    errors = []
    cleaned_row = row.copy()
    
    # Required fields for all GSTR-1 transactions
    required_fields = {
        "invoice_number": "Invoice number",
        "invoice_date": "Invoice date", 
        "invoice_value": "Invoice value",
        "place_of_supply": "Place of Supply",
        "taxable_value": "Taxable value",
        "rate": "Tax rate",
    }
    
    # Check required fields
    for field, display_name in required_fields.items():
        value = cleaned_row.get(field)
        
        # Check if missing or null
        if value is None or (isinstance(value, str) and not value.strip()):
            errors.append({
                "field": field,
                "error": f"{display_name} is required",
                "severity": "error"
            })
    
    # Safely parse numeric fields from strings
    numeric_fields = ["taxable_value", "invoice_value", "rate", "igst", "cgst", "sgst", "cess"]
    
    for field in numeric_fields:
        value = cleaned_row.get(field)
        
        if value is not None and value != "":
            try:
                # Handle percentage strings
                if isinstance(value, str):
                    value = value.strip().replace("%", "")
                    if value == "":
                        continue
                
                cleaned_row[field] = float(value)
                
            except (ValueError, TypeError) as e:
                errors.append({
                    "field": field,
                    "error": f"Could not parse {field}: {value}",
                    "severity": "error"
                })
                # Set to None if parsing fails
                cleaned_row[field] = None
    
    # Validate numeric fields have valid values (not NaN, not infinity)
    for field in numeric_fields:
        value = cleaned_row.get(field)
        if value is not None:
            import math
            if math.isnan(value) or math.isinf(value):
                errors.append({
                    "field": field,
                    "error": f"Invalid numeric value for {field}",
                    "severity": "error"
                })
                cleaned_row[field] = None
    
    # Validate tax breakup consistency
    taxable_value = cleaned_row.get("taxable_value")
    rate = cleaned_row.get("rate")
    igst = cleaned_row.get("igst", 0)
    cgst = cleaned_row.get("cgst", 0)
    sgst = cleaned_row.get("sgst", 0)
    
    if taxable_value is not None and rate is not None:
        expected_total = taxable_value * rate / 100
        
        # Check IGST
        if igst is not None and igst > 0:
            if abs(igst - expected_total) > 0.10:
                errors.append({
                    "field": "igst",
                    "error": f"IGST mismatch: expected ~{expected_total:.2f}, got {igst:.2f}",
                    "severity": "warning"
                })
        
        # Check CGST + SGST
        if cgst is not None and sgst is not None:
            expected_cgst_sgst = expected_total / 2
            actual_cgst_sgst = cgst + sgst
            if actual_cgst_sgst > 0 and abs(actual_cgst_sgst - expected_total) > 0.10:
                errors.append({
                    "field": "cgst+sgst",
                    "error": f"CGST+SGST mismatch: expected ~{expected_total:.2f}, got {actual_cgst_sgst:.2f}",
                    "severity": "warning"
                })
    
    # Log skipped rows
    if errors and row_number > 0:
        error_msgs = [e["error"] for e in errors]
        logger.warning(f"Row {row_number} ({sheet_name}): Skipped - {', '.join(error_msgs)}")
    
    # Filter to only errors (exclude warnings for row inclusion)
    error_count = sum(1 for e in errors if e["severity"] == "error")
    is_valid = error_count == 0
    
    return is_valid, errors, cleaned_row


def classify_transaction(row: Dict[str, Any], company_gstin: str = "") -> str:
    """
    Classify transaction into B2B, B2CL, B2CS, or EXP category.
    
    Args:
        row: Mapped row data
        company_gstin: Company's GSTIN for inter-state determination
        
    Returns:
        Category string
    """
    gstin = row.get("gstin", "") or ""
    pos = str(row.get("place_of_supply", "")).lower()
    invoice_value = float(row.get("invoice_value", 0))
    
    # Check for export (look for "96" or "overseas" or "other countries")
    if "96" in pos or "overseas" in pos or "other countries" in pos:
        return "EXP"
    
    # B2B if GSTIN present
    if gstin:
        return "B2B"
    
    # Extract state code from POS
    pos_code = ""
    code_match = re.match(r"^(\d{2})", pos)
    if code_match:
        pos_code = code_match.group(1)
    
    # Extract company state from GSTIN
    company_state = company_gstin[:2] if company_gstin and len(company_gstin) >= 2 else ""
    
    # Check if inter-state
    is_inter_state = company_state and pos_code and company_state != pos_code
    
    # B2CL threshold (currently Rs 2.5 lakh)
    b2cl_threshold = 250000
    
    if is_inter_state and invoice_value > b2cl_threshold:
        return "B2CL"
    
    return "B2CS"


def find_column(headers: List[str], column_type: str) -> Optional[str]:
    """
    Find the actual column header that matches a column type.
    
    Args:
        headers: List of column headers from Excel
        column_type: Type of column to find (e.g., "gstin")
        
    Returns:
        Matched column header or None
    """
    aliases = COLUMN_ALIASES.get(column_type, [])
    
    for alias in aliases:
        for header in headers:
            if header.strip().lower() == alias.lower():
                return header
    
    # Try partial matching
    for header in headers:
        header_lower = header.strip().lower()
        for alias in aliases:
            alias_lower = alias.lower()
            if alias_lower in header_lower or header_lower in alias_lower:
                return header
    
    return None


def clean_value(value: Any) -> Any:
    """
    Clean and normalize a value from Excel.
    
    Args:
        value: Raw value from Excel
        
    Returns:
        Cleaned value
    """
    if value is None:
        return None
    
    if isinstance(value, str):
        return value.strip() or None
    
    return value


def detect_header_row(worksheet, max_rows: int = 10) -> int:
    """
    Detect the row number where actual column headers start.
    
    This function handles Excel files with merged title/info rows at the top
    (common in Tally exports) where headers start from row 5.
    
    Args:
        worksheet: openpyxl worksheet object
        max_rows: Maximum rows to scan for headers (default 10)
        
    Returns:
        Row number (1-based) where headers are found, or 1 if not detected
    """
    HEADER_KEYWORDS = [
        "invoice", "gstin", "hsn", "taxable", "rate", "amount",
        "date", "value", "number", "customer", "party", "ledger",
        "tax", "cgst", "sgst", "igst", "cess", "voucher"
    ]
    
    for row_idx in range(1, min(max_rows + 1, worksheet.max_row + 1)):
        first_cell = worksheet.cell(row=row_idx, column=1).value
        if first_cell and str(first_cell).strip():
            # Check if this row contains header keywords
            row_values = [
                str(worksheet.cell(row=row_idx, column=c).value or "").lower()
                for c in range(1, min(6, worksheet.max_column + 1))
            ]
            row_text = " ".join(row_values)
            
            # Check if row has header-like content
            keyword_count = sum(1 for kw in HEADER_KEYWORDS if kw in row_text)
            if keyword_count >= 2:
                logger.debug(f"Detected header row at index {row_idx}")
                return row_idx
            
            # Also check if row doesn't look like a data row (no dates, no numbers)
            if keyword_count >= 1:
                # Check if all values in row are strings (likely headers)
                all_strings = all(
                    worksheet.cell(row=row_idx, column=c).value is None or
                    isinstance(worksheet.cell(row=row_idx, column=c).value, str)
                    for c in range(1, min(6, worksheet.max_column + 1))
                )
                if all_strings:
                    logger.debug(f"Detected header row at index {row_idx} (string-only row)")
                    return row_idx
    
    logger.debug(f"No header row detected, using default row 1")
    return 1


def stream_rows_with_header(
    worksheet,
    headers: List[str],
    header_row: int = 1,
    data_start_row: int = 2
) -> Generator[Tuple[int, Dict[str, Any]], None, None]:
    """
    Stream rows from Excel worksheet starting from a specific row.
    
    Args:
        worksheet: openpyxl worksheet object
        headers: List of column headers
        header_row: Row number where headers are located
        data_start_row: Row number where data starts
        
    Yields:
        Tuple of (row_number, row_data dictionary)
    """
    for row_idx, row in enumerate(
        worksheet.iter_rows(min_row=data_start_row, values_only=True),
        start=data_start_row
    ):
        # Skip empty rows
        if not any(cell is not None for cell in row):
            continue
            
        row_data = {}
        for idx, cell_value in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = cell_value
        
        yield row_idx, row_data


def process_excel_with_skiprows(
    file_content: bytes,
    skiprows: int = 4,
    expected_columns: Optional[List[str]] = None,
    return_period: str = "",
    company_gstin: str = ""
) -> Dict[str, Any]:
    """
    Process Excel file with skiprows parameter (e.g., Tally exports with multi-row headers).
    
    This function uses pandas for efficient row skipping when pandas is available,
    otherwise falls back to openpyxl.
    
    Args:
        file_content: Raw bytes of the Excel file
        skiprows: Number of rows to skip before reading headers (default 4 for row 5)
        expected_columns: Optional list of expected column types to validate
        return_period: Return period in MMYYYY format (for date validation)
        company_gstin: Company's GSTIN for inter-state determination
        
    Returns:
        Dictionary with:
        - "clean_data": List of valid rows
        - "errors": List of validation errors
        - "skipped_rows": List of skipped rows with reasons
        - "summary": Processing summary
    """
    start_time = time.time()
    clean_data = []
    errors = []
    skipped_rows = []
    total_rows = 0
    
    try:
        if HAS_PANDAS and pd is not None:
            # Use pandas for efficient skiprows handling
            logger.info(f"Using pandas to read Excel with skiprows={skiprows}")
            df = pd.read_excel(io.BytesIO(file_content), skiprows=skiprows)
            
            # Get headers from DataFrame columns
            headers = [str(col) for col in df.columns.tolist()]
            logger.info(f"Pandas loaded: {len(headers)} columns")
            
            # Build column mapping
            column_mapping = {}
            if expected_columns is None:
                expected_columns = list(COLUMN_ALIASES.keys())
            
            for col_type in expected_columns:
                found = find_column(headers, col_type)
                if found:
                    column_mapping[col_type] = found
            
            logger.info(f"Column mapping: {len(column_mapping)} columns mapped")
            
            # Process rows using pandas
            for row_number, (idx, row) in enumerate(df.iterrows(), start=skiprows + 2):
                # Convert row to dictionary
                row_dict = {str(col): row[col] for col in df.columns}
                
                # Map columns
                mapped_row = {}
                for std_name, actual_name in column_mapping.items():
                    if actual_name and actual_name in row_dict:
                        value = clean_value(row_dict[actual_name])
                        if value is not None:
                            mapped_row[std_name] = value
                
                # Skip empty rows
                if not mapped_row:
                    continue
                
                total_rows += 1
                
                # Classify and validate
                category = classify_transaction(mapped_row, company_gstin)
                mapped_row["_category"] = category
                
                is_valid, row_errors, cleaned_row = validate_row_for_gstr1(mapped_row, row_number, category)
                
                if row_errors:
                    error_msgs = [e["error"] for e in row_errors]
                    skipped_rows.append({
                        "row": row_number,
                        "reason": "; ".join(error_msgs),
                        "data": mapped_row
                    })
                    errors.append({"row": row_number, "errors": row_errors, "data": mapped_row})
                else:
                    clean_data.append(cleaned_row)
        else:
            # Fallback to openpyxl with manual skiprows handling
            logger.info(f"Using openpyxl to read Excel with manual skiprows={skiprows}")
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            worksheet = workbook.active
            
            # Detect header row (use skiprows if provided)
            if skiprows > 0:
                header_row = skiprows
            else:
                header_row = detect_header_row(worksheet)
            
            # Get headers from detected row
            headers = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"_col_{col}")
            
            logger.info(f"Openpyxl loaded: {len(headers)} columns from row {header_row}")
            
            # Build column mapping
            column_mapping = {}
            if expected_columns is None:
                expected_columns = list(COLUMN_ALIASES.keys())
            
            for col_type in expected_columns:
                found = find_column(headers, col_type)
                if found:
                    column_mapping[col_type] = found
            
            logger.info(f"Column mapping: {len(column_mapping)} columns mapped")
            
            # Stream rows starting from header_row + 1
            data_start_row = header_row + 1
            
            for row_number, row_data in stream_rows_with_header(
                worksheet, headers, header_row, data_start_row
            ):
                total_rows += 1
                
                # Map columns
                mapped_row = {}
                for std_name, actual_name in column_mapping.items():
                    if actual_name and actual_name in row_data:
                        value = clean_value(row_data[actual_name])
                        if value is not None:
                            mapped_row[std_name] = value
                
                # Skip empty rows
                if not mapped_row:
                    continue
                
                # Classify and validate
                category = classify_transaction(mapped_row, company_gstin)
                mapped_row["_category"] = category
                
                is_valid, row_errors, cleaned_row = validate_row_for_gstr1(mapped_row, row_number, category)
                
                if row_errors:
                    error_msgs = [e["error"] for e in row_errors]
                    skipped_rows.append({
                        "row": row_number,
                        "reason": "; ".join(error_msgs),
                        "data": mapped_row
                    })
                    errors.append({"row": row_number, "errors": row_errors, "data": mapped_row})
                else:
                    clean_data.append(cleaned_row)
            
            workbook.close()
        
        elapsed = time.time() - start_time
        
        # Build summary
        def safe_float(value, default=0.0):
            return float(value) if value is not None and value != "" else default
        
        total_taxable = sum(safe_float(row.get("taxable_value", 0)) for row in clean_data)
        total_igst = sum(safe_float(row.get("igst", 0)) for row in clean_data)
        total_cgst = sum(safe_float(row.get("cgst", 0)) for row in clean_data)
        total_sgst = sum(safe_float(row.get("sgst", 0)) for row in clean_data)
        total_cess = sum(safe_float(row.get("cess", 0)) for row in clean_data)
        
        summary = {
            "total_rows": total_rows,
            "valid_rows": len(clean_data),
            "error_rows": len(errors),
            "skipped_rows": len(skipped_rows),
            "total_taxable_value": round(total_taxable, 2),
            "total_igst": round(total_igst, 2),
            "total_cgst": round(total_cgst, 2),
            "total_sgst": round(total_sgst, 2),
            "total_cess": round(total_cess, 2),
            "processing_time_seconds": round(elapsed, 2),
        }
        
        logger.info(
            f"Excel processing completed in {elapsed:.2f}s: "
            f"{len(clean_data)} valid, {len(errors)} errors, {len(skipped_rows)} skipped"
        )
        
        return {
            "clean_data": clean_data,
            "errors": errors,
            "skipped_rows": skipped_rows,
            "summary": summary
        }
        
    except Exception as e:
        logger.exception(f"Error processing Excel file with skiprows: {str(e)}")
        return {
            "clean_data": [],
            "errors": [{"row": 0, "error": f"File processing error: {str(e)}"}],
            "skipped_rows": [],
            "summary": {"total_rows": 0, "valid_rows": 0, "error_rows": 1}
        }


def stream_rows(worksheet, headers: List[str]) -> Generator[Dict[str, Any], None, None]:
    """
    Stream rows from Excel worksheet one at a time to minimize memory usage.
    
    This generator yields rows one-by-one instead of loading all into memory,
    which is critical for large files with 10,000+ rows.
    
    Args:
        worksheet: openpyxl worksheet object
        headers: List of column headers
        
    Yields:
        Row data dictionary for each row
    """
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        # Skip empty rows
        if not any(cell is not None for cell in row):
            continue
            
        row_data = {}
        for idx, cell_value in enumerate(row):
            if idx < len(headers):
                row_data[headers[idx]] = cell_value
        
        yield row_idx, row_data


def process_excel_chunk(
    chunk: List[Dict[str, Any]],
    column_mapping: Dict[str, str],
    company_gstin: str = "",
    return_period: str = "",
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    """
    Process a chunk of rows efficiently.
    
    Preserves IRN, ack_no, ack_date, and other optional fields.
    
    Args:
        chunk: List of row data dictionaries
        column_mapping: Mapping from standard names to actual column names
        company_gstin: Company's GSTIN for inter-state determination
        return_period: Return period in MMYYYY format
        
    Returns:
        Tuple of (clean_data, errors, skipped_rows) for this chunk
    """
    clean_data = []
    errors = []
    skipped_rows = []
    
    for row_number, row_data in chunk:
        # Map columns using standard mapping
        mapped_row = map_columns(row_data, column_mapping)
        
        # Also preserve any unmapped columns (IRN, ack_no, ack_date, etc.)
        preserve_fields = [
            "irn", "ack_no", "ack_date",
            "hsn_code", "description", "quantity", "uom",
            "shipping_bill_number", "shipping_bill_date",
            "port_code", "note_number", "note_date", "note_value",
            "is_return", "is_debit_note", "reverse_charge",
            "gst_treatment", "gst_category", "is_export",
            "is_export_with_gst", "is_export_without_gst",
        ]
        
        for field in preserve_fields:
            # Check if field exists in row_data with various case combinations
            for header in row_data.keys():
                header_lower = header.strip().lower().replace("_", " ").replace("-", " ")
                field_lower = field.replace("_", " ").replace("-", " ")
                if field_lower in header_lower or header_lower in field_lower:
                    if header not in column_mapping.values():
                        value = clean_value(row_data[header])
                        if value is not None:
                            mapped_row[field] = value
                    break
        
        # Classify and validate
        category = classify_transaction(mapped_row, company_gstin)
        mapped_row["_category"] = category
        
        # Use new validation function
        is_valid, row_errors, cleaned_row = validate_row_for_gstr1(mapped_row, row_number, category)
        
        # GSTIN validation for B2B
        if mapped_row.get("gstin"):
            is_gstin_valid, error = validate_gstin(mapped_row["gstin"], is_mandatory=(category == "B2B"))
            if not is_gstin_valid:
                row_errors.append({"field": "gstin", "error": error, "severity": "error"})
        
        # POS validation
        if mapped_row.get("place_of_supply"):
            is_valid_pos, error = validate_pos(mapped_row["place_of_supply"])
            if not is_valid_pos:
                row_errors.append({"field": "place_of_supply", "error": error, "severity": "error"})
        
        # Invoice date validation
        if mapped_row.get("invoice_date"):
            is_valid_date, error, parsed_date = validate_invoice_date(mapped_row["invoice_date"])
            if not is_valid_date:
                row_errors.append({"field": "invoice_date", "error": error, "severity": "error"})
            else:
                cleaned_row["invoice_date"] = parsed_date
                if return_period:
                    is_valid_period, error = validate_return_period(parsed_date, return_period)
                    if not is_valid_period:
                        row_errors.append({"field": "invoice_date", "error": error, "severity": "warning"})
        
        # Tax rate validation
        if cleaned_row.get("rate") is not None:
            is_valid_rate, error, rate = validate_tax_rate(cleaned_row["rate"])
            if not is_valid_rate:
                row_errors.append({"field": "rate", "error": error, "severity": "error"})
            else:
                cleaned_row["rate"] = rate
        
        # Store results
        if row_errors:
            # Count errors vs warnings
            error_count = sum(1 for e in row_errors if e["severity"] == "error")
            
            error_msgs = [e["error"] for e in row_errors]
            skipped_rows.append({
                "row": row_number,
                "reason": "; ".join(error_msgs),
                "data": mapped_row
            })
            errors.append({"row": row_number, "errors": row_errors, "data": mapped_row})
        else:
            clean_data.append(cleaned_row)
    
    return clean_data, errors, skipped_rows


def map_columns(row: Dict[str, Any], column_mapping: Dict[str, str]) -> Dict[str, Any]:
    """
    Map row data to standardized column names.
    
    Args:
        row: Raw row data with column names as keys
        column_mapping: Mapping from standard names to actual column names
        
    Returns:
        Row data with standardized column names
    """
    mapped_row = {}
    
    for std_name, actual_name in column_mapping.items():
        if actual_name and actual_name in row:
            value = clean_value(row[actual_name])
            mapped_row[std_name] = value
    
    return mapped_row


def process_excel(
    file_content: bytes,
    expected_columns: Optional[List[str]] = None,
    return_period: str = "",
    company_gstin: str = "",
    skiprows: Optional[int] = None
) -> Dict[str, Any]:
    """
    Process an Excel file and validate all rows with GSTR-1 schema validations.
    
    Memory-optimized for large files with 10,000+ rows using streaming.
    Supports files with multi-row headers (e.g., Tally exports) via skiprows parameter.
    
    Args:
        file_content: Raw bytes of the Excel file
        expected_columns: Optional list of expected column types to validate
        return_period: Return period in MMYYYY format (for date validation)
        company_gstin: Company's GSTIN for inter-state determination
        skiprows: Number of rows to skip before reading headers (e.g., 4 for row 5).
                  If None, automatically detects header row position.
        
    Returns:
        Dictionary with:
        - "clean_data": List of valid rows
        - "errors": List of validation errors
        - "skipped_rows": List of skipped rows with reasons
        - "summary": Processing summary
        - "pivot_data": HSN summary from Pivot sheet (if present)
    """
    start_time = time.time()
    clean_data = []
    errors = []
    skipped_rows = []
    pivot_data = []
    total_rows = 0
    
    try:
        # If skiprows is specified, use the dedicated function
        if skiprows is not None:
            logger.info(f"Using skiprows={skiprows} for multi-row header detection")
            return process_excel_with_skiprows(
                file_content,
                skiprows=skiprows,
                expected_columns=expected_columns,
                return_period=return_period,
                company_gstin=company_gstin
            )
        
        # Load workbook from bytes - use data_only for better performance
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        worksheet = workbook.active
        
        # Detect header row position automatically
        header_row = detect_header_row(worksheet)
        
        # Get headers from detected row
        headers = []
        for cell in next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)):
            if cell is not None:
                headers.append(str(cell))
        
        logger.info(f"Excel loaded: {len(headers)} columns from row {header_row}, processing rows...")
        
        # Build column mapping
        column_mapping = {}
        if expected_columns is None:
            expected_columns = list(COLUMN_ALIASES.keys())
        
        for col_type in expected_columns:
            found = find_column(headers, col_type)
            if found:
                column_mapping[col_type] = found
        
        logger.info(f"Column mapping: {len(column_mapping)} columns mapped")
        
        # Stream rows and process in chunks for memory efficiency
        CHUNK_SIZE = 1000  # Process 1000 rows at a time
        current_chunk = []
        
        for row_number, row_data in stream_rows_with_header(
            worksheet, headers, header_row, header_row + 1
        ):
            total_rows += 1
            current_chunk.append((row_number, row_data))
            
            # Process chunk when full
            if len(current_chunk) >= CHUNK_SIZE:
                chunk_clean, chunk_errors, chunk_skipped = process_excel_chunk(
                    current_chunk, column_mapping, company_gstin, return_period
                )
                clean_data.extend(chunk_clean)
                errors.extend(chunk_errors)
                skipped_rows.extend(chunk_skipped)
                current_chunk = []
                
                # Log progress periodically
                if total_rows % 5000 == 0:
                    elapsed = time.time() - start_time
                    logger.info(f"Processed {total_rows} rows in {elapsed:.2f}s...")
        
        # Process remaining rows
        if current_chunk:
            chunk_clean, chunk_errors, chunk_skipped = process_excel_chunk(
                current_chunk, column_mapping, company_gstin, return_period
            )
            clean_data.extend(chunk_clean)
            errors.extend(chunk_errors)
            skipped_rows.extend(chunk_skipped)
        
        # Close workbook to free memory
        workbook.close()
        
        # Process Pivot sheet if present (for large files, this is a separate step)
        if "Pivot" in workbook.sheetnames:
            logger.info("Processing Pivot sheet for HSN data...")
            pivot_sheet = workbook["Pivot"]
            pivot_data = _extract_pivot_data(pivot_sheet)
        
        elapsed = time.time() - start_time
        
        # Log skipped rows summary
        if skipped_rows:
            logger.warning(f"Skipped {len(skipped_rows)} rows due to validation errors:")
            for skipped in skipped_rows[:5]:  # Log first 5
                logger.warning(f"  Row {skipped['row']}: {skipped['reason'][:100]}")
            if len(skipped_rows) > 5:
                logger.warning(f"  ... and {len(skipped_rows) - 5} more rows")
        
        # Log performance warning if slow
        if elapsed > PROCESSING_TIMEOUT_THRESHOLD:
            logger.warning(
                f"Excel processing completed in {elapsed:.2f}s for {total_rows} rows. "
                f"Average: {total_rows/elapsed:.0f} rows/sec, "
                f"Valid: {len(clean_data)}, Errors: {len(errors)}, Skipped: {len(skipped_rows)}"
            )
        else:
            logger.info(
                f"Excel processing completed in {elapsed:.2f}s: "
                f"{len(clean_data)} valid, {len(errors)} errors, {len(skipped_rows)} skipped"
            )
        
        # Build summary - ONLY from valid rows
        def safe_float(value, default=0.0):
            return float(value) if value is not None and value != "" else default
        
        total_taxable = sum(safe_float(row.get("taxable_value", 0)) for row in clean_data)
        total_igst = sum(safe_float(row.get("igst", 0)) for row in clean_data)
        total_cgst = sum(safe_float(row.get("cgst", 0)) for row in clean_data)
        total_sgst = sum(safe_float(row.get("sgst", 0)) for row in clean_data)
        total_cess = sum(safe_float(row.get("cess", 0)) for row in clean_data)
        
        summary = {
            "total_rows": total_rows,
            "valid_rows": len(clean_data),
            "error_rows": len(errors),
            "skipped_rows": len(skipped_rows),
            "total_taxable_value": round(total_taxable, 2),
            "total_igst": round(total_igst, 2),
            "total_cgst": round(total_cgst, 2),
            "total_sgst": round(total_sgst, 2),
            "total_cess": round(total_cess, 2),
            "processing_time_seconds": round(elapsed, 2),
        }
        
        result = {
            "clean_data": clean_data,
            "errors": errors,
            "skipped_rows": skipped_rows,
            "summary": summary
        }
        
        if pivot_data:
            result["pivot_data"] = pivot_data
            result["summary"]["hsn_records"] = len(pivot_data)
        
        return result
    
    except Exception as e:
        logger.exception(f"Error processing Excel file: {str(e)}")
        return {
            "clean_data": [],
            "errors": [{"row": 0, "error": f"File processing error: {str(e)}"}],
            "skipped_rows": [],
            "summary": {"total_rows": 0, "valid_rows": 0, "error_rows": 1}
        }


def _extract_pivot_data(pivot_sheet) -> List[Dict[str, Any]]:
    """
    Extract HSN data from Pivot sheet.
    
    Args:
        pivot_sheet: openpyxl worksheet object for Pivot sheet
        
    Returns:
        List of HSN records
    """
    pivot_headers = []
    for cell in next(pivot_sheet.iter_rows(min_row=1, max_row=1, values_only=True)):
        if cell is not None:
            pivot_headers.append(str(cell).strip())
    
    # Find HSN columns
    hsn_code_col = desc_col = qty_col = uom_col = None
    taxable_col = igst_col = cgst_col = sgst_col = cess_col = None
    
    for idx, header in enumerate(pivot_headers):
        header_lower = header.lower()
        if "hsn" in header_lower and hsn_code_col is None:
            hsn_code_col = idx
        elif "description" in header_lower or "desc" in header_lower:
            desc_col = idx
        elif "quantity" in header_lower or "qty" in header_lower:
            qty_col = idx
        elif "uom" in header_lower or "uqc" in header_lower:
            uom_col = idx
        elif "taxable" in header_lower:
            taxable_col = idx
        elif "integrated" in header_lower or "igst" in header_lower:
            igst_col = idx
        elif "central" in header_lower or "cgst" in header_lower:
            cgst_col = idx
        elif "state" in header_lower or "sgst" in header_lower:
            sgst_col = idx
        elif "cess" in header_lower:
            cess_col = idx
    
    # Extract data
    pivot_data = []
    for row in pivot_sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        
        hsn_record = {}
        if hsn_code_col is not None and row[hsn_code_col]:
            hsn_record["hsn_code"] = clean_value(row[hsn_code_col])
        if desc_col is not None:
            hsn_record["description"] = clean_value(row[desc_col])
        if uom_col is not None:
            hsn_record["uom"] = clean_value(row[uom_col])
        
        for col_idx, val_key in [(qty_col, "quantity"), (taxable_col, "taxable_value"),
                                   (igst_col, "igst"), (cgst_col, "cgst"),
                                   (sgst_col, "sgst"), (cess_col, "cess")]:
            if col_idx is not None and row[col_idx]:
                try:
                    hsn_record[val_key] = float(row[col_idx])
                except (ValueError, TypeError):
                    pass
        
        if hsn_record.get("hsn_code"):
            pivot_data.append(hsn_record)
    
    logger.info(f"Extracted {len(pivot_data)} HSN records from Pivot sheet")
    return pivot_data


def process_multi_sheet_excel(file_content: bytes, return_period: str = "", company_gstin: str = "") -> Dict[str, Any]:
    """
    Process Excel file with multiple pre-classified sheets (B2B, B2CL, B2CS, Export).
    
    This function handles Excel files in the GSTR-1 offline utility format
    where each sheet contains pre-classified transaction data.
    
    Args:
        file_content: Raw bytes of the Excel file
        return_period: Return period in MMYYYY format
        company_gstin: Company's GSTIN for inter-state determination
        
    Returns:
        Dictionary with clean_data and errors
    """
    start_time = time.time()
    all_clean_data = []
    all_errors = []
    skipped_rows = []  # Track skipped rows with reasons
    total_rows = 0
    
    try:
        workbook = load_workbook(io.BytesIO(file_content), data_only=True)
        
        # Known GSTR-1 sheet names
        SHEET_CATEGORIES = {
            "B2B": "B2B",
            "B2CL": "B2CL", 
            "B2CS": "B2CS",
            "Export": "EXPORT",  # IRN positional format
            "EXP": "EXPORT",
        }
        
        # Column mappings for each sheet type
        SHEET_COLUMN_MAPPINGS = {
            "B2B": {
                "GSTIN/UIN of Recipient": "gstin",
                "Receiver Name": "customer_name",
                "Invoice Number": "invoice_number",
                "Invoice date": "invoice_date",
                "Invoice Value": "invoice_value",
                "Place Of Supply": "place_of_supply",
                "Reverse Charge": "reverse_charge",
                "Invoice Type": "invoice_type",
                "E-Commerce GSTIN": "ecommerce_gstin",
                "Applicable % of Tax Rate": "diff_percent",
                "Rate": "rate",
                "Taxable Value": "taxable_value",
                "Integrated Tax Amount": "igst",
                "Central Tax Amount": "cgst",
                "State/UT Tax Amount": "sgst",
                "Cess Amount": "cess",
            },
            "B2CL": {
                "Invoice Number": "invoice_number",
                "Invoice date": "invoice_date",
                "Invoice Value": "invoice_value",
                "Place Of Supply": "place_of_supply",
                "Taxable Value": "taxable_value",
                "Rate": "rate",
                "Integrated Tax Amount": "igst",
                "Cess Amount": "cess",
                "E-Commerce GSTIN": "ecommerce_gstin",
            },
            "B2CS": {
                "Type": "type",
                "Place Of Supply": "place_of_supply",
                "Taxable Value": "taxable_value",
                "Rate": "rate",
                "Integrated Tax Amount": "igst",
                "Central Tax Amount": "cgst",
                "State/UT Tax Amount": "sgst",
                "Cess Amount": "cess",
                "E-Commerce GSTIN": "ecommerce_gstin",
            },
            "EXP": {
                "Export Type": "export_type",
                "Invoice Number": "invoice_number",
                "Invoice date": "invoice_date",
                "Invoice Value": "invoice_value",
                "Port Code": "port_code",
                "Shipping Bill Number": "shipping_bill_number",
                "Shipping Bill Date": "shipping_bill_date",
                "Taxable Value": "taxable_value",
                "Rate": "rate",
                "Integrated Tax Amount": "igst",
                "Cess Amount": "cess",
            },
            # IRN-generated format for Export sheet with special headers
            "EXPORT_IRN": {
                "State": ("state_code", str),
                "Month": (None, None),  # Skip month
                "Sl. No.": (None, None),  # Skip serial number
                "Recipient GSTIN": ("gstin", str),
                "Document Number": ("invoice_number", str),
                "Invoice Number": ("invoice_number", str),
                "Document Date": ("invoice_date", str),
                "Invoice Date": ("invoice_date", str),
                "Document Type ": (None, None),  # Trailing space
                "Document Type": (None, None),
                "Supply Type Code": (None, None),
                "Export Type": ("export_type", str),
                "Basic": ("taxable_value", float),
                # Handle headers with Indian Rupee symbol
                "Basic ": ("taxable_value", float),
                "Value": ("invoice_value", float),
                "Invoice Value": ("invoice_value", float),
                "Tax": (None, None),  # Tax is included in basic
                "Tax ": (None, None),
                "IRN": ("irn", str),
                "Status": (None, None),
                "Ack No": ("ack_no", str),
                "Acknowledgement No": ("ack_no", str),
                "Ack Date": ("ack_date", str),
                "Acknowledgement Date": ("ack_date", str),
            },
        }
        
        # Process each known sheet
        for sheet_name in workbook.sheetnames:
            if sheet_name not in SHEET_CATEGORIES:
                continue
                
            category = SHEET_CATEGORIES[sheet_name]
            worksheet = workbook[sheet_name]
            
            # Check if this is an IRN-generated Export sheet with special headers
            is_irn_export_format = False
            
            # For Export sheet, check for IRN-style headers
            if sheet_name == "Export":
                # Check first row for IRN-style headers
                first_row_headers = [str(worksheet.cell(row=1, column=c).value or "").strip() for c in range(1, 8)]
                irn_headers = ["State", "Month", "Sl. No.", "Recipient GSTIN", "Document Number", "Document Date"]
                if any(h in first_row_headers for h in irn_headers):
                    is_irn_export_format = True
                    logger.info(f"Detected IRN Export format with headers: {first_row_headers}")
            
            if is_irn_export_format:
                # Process using IRN header mapping
                header_mapping = SHEET_COLUMN_MAPPINGS.get("EXPORT_IRN", {})
                
                # Get headers from row 1
                headers = []
                for col in range(1, worksheet.max_column + 1):
                    cell_value = worksheet.cell(row=1, column=col).value
                    headers.append(str(cell_value).strip() if cell_value else f"_col_{col}")
                
                # Get all data rows (after header row 1)
                data_rows = list(worksheet.iter_rows(min_row=2, values_only=True))
                
                for row_number, row in enumerate(data_rows, start=2):
                    first_cell = row[0] if row else None
                    
                    # Skip empty rows
                    if first_cell is None or str(first_cell).strip() == "":
                        continue
                    
                    # Skip summary rows
                    if str(first_cell).lower() in ["total", "summary", "grand total"]:
                        continue
                    
                    total_rows += 1
                    
                    # Initialize row_data dictionary
                    row_data = {}
                    
                    # Map columns by header name
                    # First try exact match, then try without trailing spaces
                    for col_idx, cell_value in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            header_stripped = header.strip()
                            
                            # Try exact match first
                            if header in header_mapping:
                                field_name, field_type = header_mapping[header]
                                if field_name and cell_value is not None:
                                    try:
                                        if field_type == float:
                                            row_data[field_name] = float(cell_value)
                                        else:
                                            row_data[field_name] = str(cell_value).strip()
                                    except (ValueError, TypeError):
                                        pass
                            # Try stripped match
                            elif header_stripped in header_mapping:
                                field_name, field_type = header_mapping[header_stripped]
                                if field_name and cell_value is not None:
                                    try:
                                        if field_type == float:
                                            row_data[field_name] = float(cell_value)
                                        else:
                                            row_data[field_name] = str(cell_value).strip()
                                    except (ValueError, TypeError):
                                        pass
                            # Try without special characters (like ₹)
                            else:
                                header_clean = header_stripped.replace("₹", "").strip()
                                if header_clean in header_mapping:
                                    field_name, field_type = header_mapping[header_clean]
                                    if field_name and cell_value is not None:
                                        try:
                                            if field_type == float:
                                                row_data[field_name] = float(cell_value)
                                            else:
                                                row_data[field_name] = str(cell_value).strip()
                                        except (ValueError, TypeError):
                                            pass
                    
                    # Convert state code to POS format
                    if "state_code" in row_data:
                        state_code = str(row_data["state_code"]).upper()
                        pos_code = STATE_CODE_TO_POS.get(state_code, "96")
                        row_data["place_of_supply"] = pos_code
                    
                    # Set export type
                    export_type = row_data.get("export_type", "")
                    if export_type == "EXPWOP":
                        row_data["is_export_without_gst"] = True
                        row_data["rate"] = 0
                    elif export_type == "EXPWP":
                        row_data["is_export_with_gst"] = True
                    
                    # For exports, taxable_value = invoice_value
                    if "invoice_value" in row_data and "taxable_value" not in row_data:
                        row_data["taxable_value"] = row_data["invoice_value"]
                    
                    # Add category
                    row_data["_category"] = "EXP"
                    
                    # Validate row
                    is_valid, errors, cleaned_row = validate_row_for_gstr1(row_data, row_number, sheet_name)
                    
                    if not is_valid:
                        error_msgs = [e["error"] for e in errors]
                        skipped_rows.append({
                            "row": row_number,
                            "sheet": sheet_name,
                            "reason": "; ".join(error_msgs),
                            "data": row_data
                        })
                        all_errors.append({
                            "row": row_number, 
                            "errors": errors,
                            "data": row_data,
                            "sheet": sheet_name
                        })
                    else:
                        all_clean_data.append(cleaned_row)
                
                continue  # Skip to next sheet
            
            # Find header row - look for "Invoice" or "GSTIN" in first few rows
            header_row = 1
            for row_idx in range(1, min(6, worksheet.max_row + 1)):
                first_cell = worksheet.cell(row=row_idx, column=1).value
                if first_cell and str(first_cell).strip():
                    # Check if this row contains header keywords
                    row_values = [str(worksheet.cell(row=row_idx, column=c).value or "").lower() 
                                 for c in range(1, min(5, worksheet.max_column + 1))]
                    row_text = " ".join(row_values)
                    if "invoice" in row_text or "gstin" in row_text or "hsn" in row_text:
                        header_row = row_idx
                        break
            
            # Get headers from the detected header row
            headers = []
            for col in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=header_row, column=col).value
                if cell_value is not None:
                    headers.append(str(cell_value).strip())
                else:
                    headers.append(f"_col_{col}")
            
            # Skip if this doesn't look like a header row
            if not any(h in " ".join(headers).lower() for h in ["invoice", "gstin", "hsn", "taxable"]):
                continue
            
            # Get column mapping for this sheet
            col_mapping = SHEET_COLUMN_MAPPINGS.get(sheet_name, {})
            
            # Get all data rows (after header)
            data_rows = list(worksheet.iter_rows(min_row=header_row + 1, values_only=True))
            
            # Process data rows
            for row_number, row in enumerate(data_rows, start=header_row + 2):
                first_cell = row[0] if row else None
                
                # Skip empty rows or summary rows
                if first_cell is None or str(first_cell).strip() == "":
                    continue
                if str(first_cell).lower() in ["total", "summary", "grand total"]:
                    continue
                
                total_rows += 1
                
                # Map columns
                row_data = {}
                for col_idx, cell_value in enumerate(row):
                    if col_idx < len(headers):
                        header = headers[col_idx]
                        if header in col_mapping:
                            std_name = col_mapping[header]
                            row_data[std_name] = clean_value(cell_value)
                
                # Add category
                row_data["_category"] = category
                
                # Validate row using new validation function
                is_valid, errors, cleaned_row = validate_row_for_gstr1(row_data, row_number, sheet_name)
                
                if not is_valid:
                    # Collect error messages for logging
                    error_msgs = [e["error"] for e in errors]
                    skipped_rows.append({
                        "row": row_number,
                        "sheet": sheet_name,
                        "reason": "; ".join(error_msgs),
                        "data": row_data
                    })
                    all_errors.append({
                        "row": row_number, 
                        "errors": errors,
                        "data": row_data,
                        "sheet": sheet_name
                    })
                else:
                    all_clean_data.append(cleaned_row)
        
        workbook.close()
        
        elapsed = time.time() - start_time
        
        # Log skipped rows summary
        if skipped_rows:
            logger.warning(f"Skipped {len(skipped_rows)} rows due to validation errors:")
            for skipped in skipped_rows[:5]:  # Log first 5
                logger.warning(f"  Row {skipped['row']} ({skipped['sheet']}): {skipped['reason'][:100]}")
            if len(skipped_rows) > 5:
                logger.warning(f"  ... and {len(skipped_rows) - 5} more rows")
        
        logger.info(f"Multi-sheet Excel processed in {elapsed:.2f}s: {len(all_clean_data)} valid, {len(all_errors)} errors, {len(skipped_rows)} skipped")
        
        # Build summary - ONLY from valid rows
        def safe_float(value, default=0.0):
            return float(value) if value is not None and value != "" else default
        
        total_taxable = sum(safe_float(row.get("taxable_value", 0)) for row in all_clean_data)
        total_igst = sum(safe_float(row.get("igst", 0)) for row in all_clean_data)
        total_cgst = sum(safe_float(row.get("cgst", 0)) for row in all_clean_data)
        total_sgst = sum(safe_float(row.get("sgst", 0)) for row in all_clean_data)
        total_cess = sum(safe_float(row.get("cess", 0)) for row in all_clean_data)
        
        summary = {
            "total_rows": total_rows,
            "valid_rows": len(all_clean_data),
            "error_rows": len(all_errors),
            "skipped_rows": len(skipped_rows),
            "total_taxable_value": round(total_taxable, 2),
            "total_igst": round(total_igst, 2),
            "total_cgst": round(total_cgst, 2),
            "total_sgst": round(total_sgst, 2),
            "total_cess": round(total_cess, 2),
            "processing_time_seconds": round(elapsed, 2),
        }
        
        return {
            "clean_data": all_clean_data,
            "errors": all_errors,
            "skipped_rows": skipped_rows,
            "summary": summary
        }
        
    except Exception as e:
        logger.exception(f"Error processing multi-sheet Excel: {str(e)}")
        return {
            "clean_data": [],
            "errors": [{"row": 0, "error": f"File processing error: {str(e)}"}],
            "skipped_rows": [],
            "summary": {"total_rows": 0, "valid_rows": 0, "error_rows": 1}
        }


def process_gstr1_excel(file_content: Any, skiprows: Optional[int] = None) -> Dict[str, Any]:
    """
    Process Excel file for GSTR-1 format.
    
    Automatically detects multi-sheet format (GSTR-1 offline utility format)
    and processes each sheet accordingly.
    
    Supports files with multi-row headers (e.g., Tally exports) via skiprows parameter.
    
    Args:
        file_content: Raw bytes of the Excel file OR file path (str)
        skiprows: Number of rows to skip before reading headers (e.g., 4 for row 5).
                  If None, automatically detects header row position.
        
    Returns:
        Dictionary with GSTR-1 structured data including:
        - clean_data: List of valid rows
        - errors: List of validation errors
        - skipped_rows: List of skipped rows with reasons
        - summary: Processing summary with totals calculated from valid rows only
    """
    # If file_content is a string (file path), read the file
    if isinstance(file_content, str):
        with open(file_content, "rb") as f:
            file_content = f.read()
    
    # If skiprows is specified, use process_excel with skiprows
    if skiprows is not None:
        logger.info(f"Using skiprows={skiprows} for GSTR-1 processing")
        result = process_excel(
            file_content,
            skiprows=skiprows
        )
    else:
        # Check if it's a multi-sheet Excel file (GSTR-1 offline utility format)
        try:
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            sheet_names = workbook.sheetnames
            workbook.close()
            
            # Check if file has known GSTR-1 sheet names
            KNOWN_SHEETS = {"B2B", "B2CL", "B2CS", "Export", "EXP", "CDNR", "CDNUR", "HSN", "Docs"}
            has_gstr1_sheets = any(sheet in KNOWN_SHEETS for sheet in sheet_names)
            
            if has_gstr1_sheets:
                logger.info(f"Detected multi-sheet GSTR-1 format with sheets: {sheet_names}")
                result = process_multi_sheet_excel(file_content)
            else:
                result = process_excel(file_content)
        except Exception as e:
            logger.exception(f"Error detecting Excel format: {str(e)}")
            result = process_excel(file_content)
    
    if result.get("errors"):
        logger.warning(f"Found {len(result['errors'])} validation errors in Excel file")
    
    # Handle skipped rows
    skipped_rows = result.get("skipped_rows", [])
    if skipped_rows:
        logger.warning(f"Skipped {len(skipped_rows)} rows due to validation errors")
    
    # Transform to GSTR-1 format - use summary from result (already calculated from valid rows)
    gstr1_data = {
        "b2b": [],
        "b2cl": [],
        "b2cs": [],
        "export": [],
        "cdnr": [],
        "cdnur": [],
        "summary": result.get("summary", {})
    }
    
    for row in result["clean_data"]:
        category = row.get("_category", "")
        gstin = row.get("gstin", "") or ""
        
        # Use pre-classified category if available
        if category == "B2B":
            gstr1_data["b2b"].append({
                "gstin": row.get("gstin", ""),
                "customer_name": row.get("customer_name", ""),
                "invoice_number": row.get("invoice_number", ""),
                "invoice_date": row.get("invoice_date", ""),
                "invoice_value": row.get("invoice_value", 0),
                "place_of_supply": row.get("place_of_supply", ""),
                "taxable_value": row.get("taxable_value", 0),
                "rate": row.get("rate", 0),
                "cgst": row.get("cgst", 0),
                "sgst": row.get("sgst", 0),
                "igst": row.get("igst", 0),
                "cess": row.get("cess", 0),
            })
        elif category == "B2CL":
            gstr1_data["b2cl"].append({
                "invoice_number": row.get("invoice_number", ""),
                "invoice_date": row.get("invoice_date", ""),
                "invoice_value": row.get("invoice_value", 0),
                "place_of_supply": row.get("place_of_supply", ""),
                "taxable_value": row.get("taxable_value", 0),
                "rate": row.get("rate", 0),
                "igst": row.get("igst", 0),
                "cess": row.get("cess", 0),
            })
        elif category == "B2CS":
            gstr1_data["b2cs"].append({
                "invoice_date": row.get("invoice_date", ""),
                "invoice_value": row.get("invoice_value", 0),
                "place_of_supply": row.get("place_of_supply", ""),
                "taxable_value": row.get("taxable_value", 0),
                "rate": row.get("rate", 0),
                "igst": row.get("igst", 0),
                "cess": row.get("cess", 0),
            })
        elif category == "EXP":
            gstr1_data["export"].append({
                "invoice_number": row.get("invoice_number", ""),
                "invoice_date": row.get("invoice_date", ""),
                "invoice_value": row.get("invoice_value", 0),
                "shipping_port": row.get("port_code", ""),
                "taxable_value": row.get("taxable_value", 0),
                "rate": row.get("rate", 0),
                "igst": row.get("igst", 0),
                "cess": row.get("cess", 0),
            })
    
    logger.info(
        f"GSTR-1 data extracted: "
        f"{len(gstr1_data['b2b'])} B2B, "
        f"{len(gstr1_data['b2cl'])} B2CL, "
        f"{len(gstr1_data['b2cs'])} B2CS, "
        f"{len(gstr1_data['export'])} Export"
    )
    
    return {
        **gstr1_data,
        "clean_data": result["clean_data"],
        "errors": result.get("errors", []),
        "skipped_rows": skipped_rows,
        "validation_summary": {
            "errors": result.get("errors", []),
            "skipped_rows": skipped_rows,
            "warnings": [],
        }
    }
