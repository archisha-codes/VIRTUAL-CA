from openpyxl import load_workbook
wb = load_workbook('gstr1_122025.xlsx')
ws = wb['B2B']
print('B2B rows:', ws.max_row)
print('B2B columns:', ws.max_column)
# Show first few rows
for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
    print([str(c)[:15] if c else '' for c in row])
